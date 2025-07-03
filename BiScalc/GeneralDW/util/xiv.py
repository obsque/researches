import copy
import openpyxl
from enum import *

class ArmType(Enum):
    a수호자 = '1수호자'
    b학살자 = '2학살자'
    c타격대 = '3타격대'
    d정찰대 = '4정찰대'
    e유격대 = '5유격대'
    f마술사 = '6마술사'
    g치유사 = '7치유사'
    pass
class AccType(Enum):
    a수호자 = '1수호자'
    b공격대 = '2공격대'
    c유격대 = '3유격대'
    d마술사 = '4마술사'
    e치유사 = '5치유사'
    pass
JOB = {
    'PLD': [ ArmType.a수호자, ArmType.a수호자, AccType.a수호자 ],
    'WAR': [ ArmType.a수호자, ArmType.a수호자, AccType.a수호자 ],
    'DRK': [ ArmType.a수호자, ArmType.a수호자, AccType.a수호자 ],
    'GNB': [ ArmType.a수호자, ArmType.a수호자, AccType.a수호자 ],
    'DRG': [ 'DRG', ArmType.b학살자, AccType.b공격대 ],
    'RPR': [ 'RPR', ArmType.b학살자, AccType.b공격대 ],
    'MNK': [ 'MNK', ArmType.c타격대, AccType.b공격대 ],
    'SAM': [ 'SAM', ArmType.c타격대, AccType.b공격대 ],
    'NIN': [ 'NIN', ArmType.d정찰대, AccType.c유격대 ],
    'VPR': [ 'VPR', ArmType.d정찰대, AccType.c유격대 ],
    'BRD': [ 'BRD', ArmType.e유격대, AccType.c유격대 ],
    'MCH': [ 'MCH', ArmType.e유격대, AccType.c유격대 ],
    'DNC': [ 'DNC', ArmType.e유격대, AccType.c유격대 ],
    'BLM': [ 'BLM', ArmType.f마술사, AccType.d마술사 ],
    'SMN': [ 'SMN', ArmType.f마술사, AccType.d마술사 ],
    'RDM': [ 'RDM', ArmType.f마술사, AccType.d마술사 ],
    'PCT': [ 'PCT', ArmType.f마술사, AccType.d마술사 ],
    'WHM': [ 'WHM', ArmType.g치유사, AccType.e치유사 ],
    'SCH': [ 'SCH', ArmType.g치유사, AccType.e치유사 ],
    'AST': [ 'AST', ArmType.g치유사, AccType.e치유사 ],
    'SGE': [ 'SGE', ArmType.g치유사, AccType.e치유사 ],
}
class Slot(Enum):
    w무기 = '0무기'
    h머리 = '1머리'
    b몸통 = '2몸통'
    a손 = '3손'
    l다리 = '4다리'
    f발 = '5발'
    e귀걸이 = 'u귀걸이'
    n목걸이 = 'v목걸이'
    r팔찌 = 'w팔찌'
    i반지 = 'x반지'
    pass

# iequips = ['weapon', 'head', 'body', 'hands', 'legs', 'feet',
#            'earrings', 'necklace', 'bracelets', 'ring1', 'ring2', 'food']
iequips = ['0무기', '1머리', '2몸통', '3손', '4다리', '5발',
           'u귀', 'v목', 'w팔', 'x반지', 'food']

istats = ['attr'
        #   , 'vital'
          , 'dh', 'crit', 'det', 'sks', 'tncpt'
          , 'matA', 'matB'
        #   , 'Elemental', 'haste'
          ] #, 'materia']
eStats = {k: v for v, k in enumerate(istats)}  # Reverse mapping for easy lookup

class ITEMTYPE(Enum):
    weapon = 'W'
    head = 'B'
    body = 'A'
    hands = 'B'
    legs = 'A'
    feet = 'B'
    earrings = 'C'
    necklace = 'C'
    bracelets = 'C'
    ring = 'C'
    etc = '0'

class Specs:
    def __init__(self, attr=0, vital=0, dh=0, crit=0, det=0,
                 sks=0, tncpt=0, elemental=0, haste=0):
        self.attr = attr
        self.vital = vital
        self.dh = dh
        self.crt = crit
        self.det = det
        self.sks = sks
        self.tncpt = tncpt
        self.matA = 0
        self.matB = 0
        # self.elemental = elemental
        # self.haste = haste

# class ITEM(Specs):
class ITEM():
    def __init__(self): # -> None:
        # super().__init__()
        self.obtain = 0
        self.type = 0
        self.slot = 0
        self.specs = {x:0 for x in istats}  # Initialize all stats to 0
        pass

def getItemsetsFromDB(path):
    wb = openpyxl.load_workbook(path)

    # sheet = wb['DB']
    sheet = wb[wb.sheetnames[0]]  # Assuming the first sheet is the one we want
    # mr = sheet.max_row
    # mc = sheet.max_column

    # Item Data
    item_dict = []
    headers = [cell.value for cell in sheet[1]]  # Get headers from the first row
    # print(f'{headers}')
    ### ['부위', '직군', '획득', '주', '극', '의', '시', '직', '굴/신']
    temp = {}
    # temp = ITEM()
    for row in sheet.iter_rows(2, values_only=True):
        # print(row)
        row_dict = {headers[i]: row[i] for i in range(len(headers))}
        row_dict[1] = row[1] if row[1] is not None else temp[1]
        item_dict.append(row_dict)
        temp = row_dict
        # temp.obtain = row[0]
        # temp.slot = row[1]
        # temp.type = row[2] if row[2] is not None else temp.type
        # temp.specs = {istats[i]: row[i+3] for i in range(6)}
        # item_dict.append(copy.copy(temp))

    return item_dict



def test():

    pass

if __name__ == "__main__":
    test()
    sample = '''부위,직군,획득,주,극,의,시,직,굴/신,matA,matB
1머리,1수호자,제작,392,241,,,,169,3,2
1머리,1수호자,일반,392,169,241,,,,2,
1머리,1수호자,석판,413,246,,,,172,2,
1머리,1수호자,보강,435,251,,,,176,2,
1머리,1수호자,레이드,435,251,176,,,,2,
2몸통,1수호자,제작,622,268,383,,,,3,2
2몸통,1수호자,일반,622,,383,,,268,2,
2몸통,1수호자,석판,657,391,274,,,,2,
2몸통,1수호자,보강,690,399,279,,,,2,
2몸통,1수호자,레이드,690,,399,,,279,2,
3손,1수호자,제작,392,241,,,,169,3,2
3손,1수호자,일반,392,,169,241,,,2,
3손,1수호자,석판,413,,172,,,246,2,
'''
    # Convert the multiline string 'sample' to a list of lists
    sample_lines = [line.strip() for line in sample.strip().split('\n') if line.strip()]
    sheet = [row.split(',') for row in sample_lines]

