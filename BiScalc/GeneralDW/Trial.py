import os
import openpyxl
import pandas as pd
from datetime import datetime
import time
import itertools

from util.cdhsp import *
import io
from util.xiv import *  # ITEM, ITEMTYPE

class ItemSet:
    def __init__(self, level, job):
        self.job = job
        self.level = level
        # self.calc = StatsCalc(level)
        self.equips = { x: [] for x in iequips }  # Initialize all equipment slots with empty lists
        pass
    pass

base = ItemSet(100, 'PLD')
### Load the Excel file
path = r'E:\Users\i\Documents\My Games\FINAL FANTASY XIV - KOREA\docs\7황금'
file = r'760items.xlsx'
file_path = os.path.join(path, file)
print(f'\nLoading {file} from {path}...')

def getItemsFromDB(file_path, filter=None):
    # item_dict = []
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[wb.sheetnames[0]]  # Assuming the first sheet is the one we want
    ### Item Data
    headers = [cell.value for cell in sheet[1]]  # Get headers from the first row
    # print(f'{headers}')
    ### ['부위', '직군', '획득', '주', '극', '의', '시', '직', '굴/신']
    temp = {}
    for row in sheet.iter_rows(2, values_only=True):
        ### JOB 착용가능 type 직군 확인
        #부위
        if row[0] == 'food':
            pass
        #직군
        elif row[1] not in [j.value for j in JOB[base.job]]:
            continue
        ### 획득: 출발 / 최종
        elif row[2] not in ['제작', '일반', '석판', '토벌']:
            continue
        idata = {headers[i]: row[i] for i in range(2, len(headers))}
        idata[headers[1]] = row[1] if row[1] is not None else temp[headers[1]]
        if row[0] not in base.equips:
            base.equips[row[0]] = [idata]
        else:
            base.equips[row[0]].append(idata)
        # item_dict.append(idata)
        temp = idata
        pass
    # return item_dict

getItemsFromDB(file_path)
# print(f'{base.job} - {len(item_dict)} items.')
print(len(base.equips))


### Calculation

### Week 1
TombMax = 900
TombCosts = {'W':900, 'A':825, 'B':495, 'C':375}

stat_base = {k: 0 for v,k in enumerate(istats)}

for i in range(len(base.equips)):
    key = list(base.equips.keys())[i]
    for s in base.equips[key]:
        istat = {
            'attr' : 0 if '주' == None else s['주'],
            'dh' : 0 if '직' == None else s['직'],
            'crit' : 0 if '극' == None else s['극'],
            'det' : 0 if '의' == None else s['의'],
            'sks' : 0 if '시' == None else s['시'],
            'tncpt' : 0 if '굴/신' == None else s['굴/신'],
            'matA' : 0 if 'matA' == None else s['matA'],
            'matB' : 0 if 'matA' == None else s['matB']
        }

        istat_total = {k: stat_base.get(k, 0) + int(istat.get(k, 0) or 0) for k in stat_base}
        print(istat_total)
    pass
