import openpyxl
from enum import Enum

iequips = ['weapon', 'head', 'body', 'hands', 'legs', 'feet',
           'earrings', 'necklace', 'bracelets', 'ring1', 'ring2', 'food']

istats = ['attr', 'vital', 'dh', 'crit', 'det',
          'sks', 'tncpt', 'Elemental', 'haste'] #, 'materia']
eStats = {'attr': 0, 'vital': 1,
          'dh': 2, 'crit': 3, 'det': 4, 'sks': 5, 'tncpt': 6,
          'Elemental': 7, 'haste': 8,
          'materia': 9}

class ITEMTYPE(Enum):
    weapon = 'W'
    head = 'B'
    body = 'A'
    hands = 'B'
    legs = 'A'
    feet = 'B'
    earrings = 'C'
    necklace = 'C'
    bracelets = 'C'
    ring = 'C'

class ITEM():
    def __init__(self, type, ): # -> None:
        self.type
        self.attr
        self.vital
        self.CRT
        self.DET
        self.DH
        self.SKS
        self.TenPie
        self.haste
        pass


def getItemsetsFromDB(path):
    wb = openpyxl.load_workbook(path)

    # sheet = wb['DB']
    sheet = wb[0]  # Assuming the first sheet is the one we want
    # mr = sheet.max_row
    # mc = sheet.max_column

    # Item Data
    item_dict = {}
    for row in sheet.iter_rows(2, values_only=True):
        # print(row)
        # (직군, 부위, name, Attr, vital, 직격, 극대, 의지, 시속, 불굴신앙, 속성, Haste, 마테)
        직군 = row[0]
        부위 = row[1]
        name = row[2]

        if 직군 not in (item_dict):
            item_dict[직군] = {}
            pass

        if 부위 not in (item_dict[직군]):
            item_dict[직군][부위] = {}
            pass

        buffer = [0 if cell is None else cell for cell in row[3:]]

        item_dict[직군][부위][name] = buffer  # .copy()

        # item_dict[직군][부위][name] = [Attr, vital, 직격, 극대, 의지, 시속, 불굴, 속성, Haste, 마테[1:5]]
    return item_dict



def test():

    pass

if __name__ == "__main__":
	test()
