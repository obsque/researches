import openpyxl
import pandas as pd
from datetime import datetime
from item import *  # ITEM, ITEMTYPE
from cdhsp import *
from calcs import *

path = "E:\GIT\eureka-bis\items.xlsx"
wb = openpyxl.load_workbook(path)
# wb = openpyxl.load_workbook('E:\\Users\\i\\Documents\\My Games\\FINAL FANTASY XIV - KOREA\\docs\\eurekaBiS.xlsx')

# sheet = wb.get_sheet_by_name('DB')
sheet = wb['DB']
mr = sheet.max_row
mc = sheet.max_column

# Item Data
item_dict = {}
for row in sheet.iter_rows(2, values_only=True):  # (2, 20, values_only=True):
    # print(row)
    # (직군, 부위, name, Attr, vital, 직격, 극대, 의지, 시속, 불굴신앙, 속성, Haste, 마테)
    직군 = row[0]
    부위 = row[1]
    name = row[2]

    if 직군 not in (item_dict):
        item_dict[직군] = {}
        pass

    if 부위 not in (item_dict[직군]):
        item_dict[직군][부위] = {}
        pass

    buffer = list(row[3:])

    item_dict[직군][부위][name] = buffer  # .copy()

    # item_dict[직군][부위][name] = [Attr, vital, 직격, 극대, 의지, 시속, 불굴, 속성, Haste, 마테[1:5]]

# ItemSet Data
jobs = {
    'TANK' : [0, 0, 80, 80, 80, 80],
    'HEAL' : [6, 4, 0, 0, 0, 0],
    'SAM' : [2, 1, 60, 84, 60, 108],
    'BRD' : [4, 2, 78, 114, 0, 114],
    'SMN' : [5, 3, 114, 72, 36, 54],
}
job = 'TANK'

LEVEL = 70
i1 = jobs[job][0]
i2 = jobs[job][1]
w_crit = jobs[job][2]
w_det = jobs[job][3]
w_sks = jobs[job][4]
w_dht = jobs[job][5]
class1 = ['수호자', '학살자', '타격대', '정찰대', '유격대', '마술사', '치유사']
class2 = ['수호자', '공격대', '유격대', '마술사', '치유사']
직군 = class1[i1]
직군2 = class2[i2]

itemset = {}
# 고정 부위
itemset['hands'] = item_dict[직군]['손']  # ['+2']
itemset['legs'] = item_dict[직군]['다리']  # ['+2']
itemset['feet'] = item_dict[직군]['발']  # ['+2']
# 가변 부위
itemset['weapon'] = {'base': [112, 111, 0, 0, 0, 0, 0, 348, 0]}
itemset['head'] = item_dict[직군]['머리']
itemset['body'] = item_dict[직군]['몸통']
# 직군2 = '유격대'
# 직군2 = '마술사'
itemset['earrings'] = item_dict['Sync']['귀']
itemset['necklace'] = item_dict['Sync']['목']
itemset['bracelets'] = item_dict['Sync']['팔']
itemset['ring1'] = item_dict['Sync']['반지'].copy()
itemset['ring2'] = item_dict['Sync']['반지'].copy()
itemset['earrings'].update(item_dict[직군2]['귀'])
itemset['necklace'].update(item_dict[직군2]['목'])
itemset['bracelets'].update(item_dict[직군2]['팔'])
itemset['ring1'].update(item_dict[직군2]['반지'])
itemset['ring2'].update(item_dict[직군2]['반지'])
# add haste items
itemset['head'].update(item_dict['Haste']['머리'])
itemset['earrings'].update(item_dict['Haste']['귀'])
itemset['ring1'].update(item_dict['Haste']['반지'])


def dh2tnc(statvalues):
    if 직군 == '수호자' or 직군 == '치유사':
        temp = statvalues[eStats['dh']]
        statvalues[eStats['dh']] = None
        statvalues[eStats['tncpt']] = temp

# [1:15](김융털) 극759 의468 직 400
# [1:15](김융털) 기982
# STATS
base_stats = {}
base_stats['attr'] = LevelMod[LEVEL]['attr']
base_stats['vital'] = LevelMod[LEVEL]['vital']
base_stats['dh'] = LevelMod[LEVEL]['main']
base_stats['crit'] = LevelMod[LEVEL]['main']
base_stats['det'] = LevelMod[LEVEL]['main']
base_stats['sks'] = LevelMod[LEVEL]['main']
base_stats['tncpt'] = LevelMod[LEVEL]['main']
base_stats['Elemental'] = 3558
base_stats['haste'] = 0

# OUTPUT
expDmg = 0.0
BiS_stats = base_stats.copy()
BiS = {}

##########

# 장비 SET
equips = {}
# 고정 장비
# equips['hands'] = item_dict[직군]['손']  # ['+2']
# equips['legs'] = item_dict[직군]['다리']  # ['+2']
# equips['feet'] = item_dict[직군]['발']  # ['+2']


def compareBiS(tmp_stats, in_equips): #, expDmg, BiS_stats, BiS):
    global base_stats
    tmp_stats = base_stats.copy()
    for name, value in equips.items():
        for stats in value.values(): # range(0, len(sum)):
            for i in range(0, 9): # len(stats)):
                # if istats[i] in istats:
                    cellvalue = stats[i]
                    if cellvalue != None:
                        tmp_stats[istats[i]] += cellvalue

    result = ExpDmgSum(tmp_stats['attr'], tmp_stats['dh'], tmp_stats['crit'],
                       tmp_stats['det'], tmp_stats['sks'], tmp_stats['tncpt'], tmp_stats['haste'])
    global expDmg, BiS_stats, BiS

    if result > expDmg:
        expDmg = result
        BiS = in_equips.copy()
        BiS_stats.update(tmp_stats)

        print('{:8.3f}'.format(expDmg))
        # print(BiS)
        for id in iequips:
            if id in BiS:
                print(f'{id} : {BiS[id]}')
        print(BiS_stats)

def calcPart(part, in_stats, NextChain=None, *args, **kwargs):
    part_stat = in_stats.copy()
    for partname, value in itemset[part].items():
        equips[part] = {partname: value}
        # addStats(part_stat, value)

        if NextChain != None:
            NextChain(part_stat, *args, **kwargs)
            pass
        # else:
            # compareBiS(part_stat, equips, expDmg, BiS_stats, BiS)


def BodyPart(in_stats, NextChain, *args, **kwargs):
    part = 'body'

    for body, value in itemset[part].items():
        equips[part] = {body: value}
        part_stat = in_stats.copy()

        if body != '기린' and body != '주홍':
            # addStats(part_stat, value)
            NextChain(part_stat, *args, **kwargs)
        else:
            if body == '주홍':
                # addStats(part_stat, -equips['head'].value().value())
                equips['head'].clear()
            for mb1 in range(2, 6):
                value31 = value.copy()
                if AddMateria(value31, mb1, 9) == False:
                    continue
                for mb2 in range(2, 6):
                    value32 = value31.copy()
                    if AddMateria(value32, mb2, 10) == False:
                        continue
                    for mb3 in range(2, 6):
                        value33 = value32.copy()
                        if AddMateria(value33, mb3, 11) == False:
                            continue
                        for mb4 in range(2, 6):
                            value34 = value33.copy()
                            if AddMateria(value34, mb4, 12) == False:
                                continue
                            for mb5 in range(2, 6):
                                value35 = value34.copy()
                                if AddMateria(value35, mb5, 13) == False:
                                    continue
                                equips[part] = {body: value35}
                                # part_stat = in_stats.copy()
                                # addStats(part_stat, value35)
                                NextChain(part_stat, *args, **kwargs)

# For Accessories

def AccPart(part, in_stats, NextChain=None, *args, **kwargs):
    for partname, value in itemset[part].items():
        equips[part] = {partname: value}

        if partname != '제왕비취':
            if NextChain == None:
                part_stat = in_stats.copy()
                # addStats(part_stat, value)
                compareBiS(part_stat, equips) #, expDmg, BiS_stats, BiS)
            else:
                part_stat = in_stats.copy()
                # addStats(part_stat, value)
                NextChain(part_stat, *args, **kwargs)
                pass
        else:  # partname == '제왕비취':
            for mb1 in range(2, 6):
                value1 = value.copy()
                if AddMateria(value1, mb1, 9, 16, 55) == False:
                    continue
                for mb2 in range(2, 6):
                    value2 = value1.copy()
                    if AddMateria(value2, mb2, 10, 16, 55) == False:
                        continue
                    for mb3 in range(2, 6):
                        value3 = value2.copy()
                        if AddMateria(value3, mb3, 11, 8, 55) == False:
                            continue
                        for mb4 in range(2, 6):
                            value4 = value3.copy()
                            if AddMateria(value4, mb4, 12, 8, 55) == False:
                                continue
                            for mb5 in range(2, 6):
                                value5 = value4.copy()
                                if AddMateria(value5, mb5, 13, 8, 55) == False:
                                    continue
                                equips[part] = {partname: value5}

                                if NextChain == None:
                                    part_stat = in_stats.copy()
                                    # addStats(part_stat, value5)
                                    compareBiS(part_stat, equips) #, expDmg, BiS_stats, BiS)
                                else:
                                    part_stat = in_stats.copy()
                                    # addStats(part_stat, value5)
                                    NextChain(part_stat, *args, **kwargs)
                                    pass


# 가변 부위
def Ring2Chain(in_stats):
    AccPart('ring2', in_stats)

def Ring1Chain(in_stats):
    AccPart('ring1', in_stats, Ring2Chain)

def BraceletsChain(in_stats):
    print(f"@@@@@@ bracelets @@@@@@  {datetime.now()}")
    AccPart('bracelets', in_stats, Ring1Chain)

def NecklaceChain(in_stats):
    print(f"@@@@@ 'necklace @@@@@")
    AccPart('necklace', in_stats, BraceletsChain)


def EarringsChain(in_stats):
    print(f"@@@ earrings @@@")
    AccPart('earrings', in_stats, NecklaceChain)

def BodyChain(in_stats):
    BodyPart(in_stats, EarringsChain)

def HeadChain(in_stats):
    calcPart('head', in_stats, BodyChain)

# 고정 부위
def FeetChain(in_stats):
    AccPart('feet', in_stats, HeadChain)

def LegsChain(in_stats):
    calcPart('legs', in_stats, FeetChain)

def HandsChain(in_stats):
    calcPart('hands', in_stats, LegsChain)


def calcPart(part, in_stats, NextChain=None, *args, **kwargs):
    part_stat = in_stats.copy()
    for partname, value in itemset[part].items():
        equips[part] = {partname: value}
        addStats(part_stat, value)

        if NextChain == None:
            compareBiS(part_stat, equips) #, expDmg, BiS_stats, BiS)
        else:
            NextChain(part_stat, *args, **kwargs)
            pass

# 무기


def WeaponPart(in_stats, dh=0, crit=0, det=0, sks=0, NextChain=None, *args, **kwargs):
    part_stats = in_stats.copy()

    for weapon, w_value in itemset['weapon'].items():
        # statlist = [36, 54, 80, 96, 114]
        # # for stat1 in statlist:
        w_value[eStats['dh']] = min(dh, 114)
        w_value[eStats['crit']] = min(crit, 114)
        w_value[eStats['det']] = min(det, 114)
        w_value[eStats['sks']] = min(sks, 114)
        # w_value[eStats['tncpt']] = 0

    dh2tnc(w_value)
    equips['weapon'] = {job: w_value}
    addStats(part_stats, w_value)
    NextChain(part_stats, *args, **kwargs)

##########


WeaponPart(base_stats, w_dht, w_crit, w_det, w_sks, HandsChain)

# R E S U L T

print('{:8.3f}'.format(expDmg))
print(BiS_stats)
# print(BiS)
for id in iequips:
    print(f'{id} : {BiS[id]}')
