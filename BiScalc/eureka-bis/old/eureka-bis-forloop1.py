import openpyxl
import pandas as pd
from item import *  # ITEM, ITEMTYPE
from cdhsp import *

# import common as c

path = "E:\GIT\eureka-bis\items.xlsx"
wb = openpyxl.load_workbook(path)
# wb = openpyxl.load_workbook('E:\\Users\\i\\Documents\\My Games\\FINAL FANTASY XIV - KOREA\\docs\\eurekaBiS.xlsx')

# sheet = wb.get_sheet_by_name('DB')
sheet = wb['DB']
mr = sheet.max_row
mc = sheet.max_column

iequips = ['weapon', 'head', 'body', 'hands', 'legs', 'feet',
           'earrings', 'necklace', 'bracelets', 'ring1', 'ring2']

istats = ['attr', 'vital', 'dh', 'crit', 'det',
          'sks', 'tncpt', 'Elemental', 'haste'] #, 'materia']
eStats = {'attr': 0, 'vital': 1,
          'dh': 2, 'crit': 3, 'det': 4, 'sks': 5, 'tncpt': 6,
          'Elemental': 7, 'haste': 8,
          'materia': 9}

# Item Data
item_dict = {}
for row in sheet.iter_rows(2, values_only=True):  # (2, 20, values_only=True):
    # print(row)
    # (직군, 부위, name, Attr, vital, 직격, 극대, 의지, 시속, 불굴신앙, 속성, Haste, 마테)
    직군 = row[0]
    부위 = row[1]
    name = row[2]

    if 직군 not in (item_dict):
        item_dict[직군] = {}
        pass

    if 부위 not in (item_dict[직군]):
        item_dict[직군][부위] = {}
        pass

    buffer = list(row[3:])

    item_dict[직군][부위][name] = buffer #.copy()
        
    # item_dict[직군][부위][name] = [Attr, vital, 직격, 극대, 의지, 시속, 불굴, 속성, Haste, 마테[1:5]]

# ItemSet Data
LEVEL = 70
# 직군 = '수호자'
# 직군 = '유격대'
직군 = '마술사'
itemset = {}
# 고정 부위
# itemset['hands'] = item_dict[직군]['손']  # ['+2']
# itemset['legs'] = item_dict[직군]['다리']  # ['+2']
# itemset['feet'] = item_dict[직군]['발']  # ['+2']
# 가변 부위
itemset['weapon'] = {'base': [112, 111, 0, 0, 0, 0, 0, 348, 0]}
itemset['head'] = item_dict[직군]['머리']
itemset['body'] = item_dict[직군]['몸통']
# 직군 = '수호자'
# 직군 = '유격대'
직군 = '마술사'
itemset['earrings'] = item_dict['Sync']['귀']
itemset['necklace'] = item_dict['Sync']['목']
itemset['bracelets'] = item_dict['Sync']['팔']
itemset['ring1'] = item_dict['Sync']['반지'].copy()
itemset['ring2'] = item_dict['Sync']['반지'].copy()
itemset['earrings'].update(item_dict[직군]['귀'])
itemset['necklace'].update(item_dict[직군]['목'])
itemset['bracelets'].update(item_dict[직군]['팔'])
itemset['ring1'].update(item_dict[직군]['반지'])
itemset['ring2'].update(item_dict[직군]['반지'])
# add haste items
itemset['head'].update(item_dict['Haste']['머리'])
itemset['earrings'].update(item_dict['Haste']['귀'])
# itemset['ring1'].update(item_dict['Haste']['반지'])

# STATS
base_stats = {}
base_stats['attr'] = LevelMod[LEVEL]['attr']
base_stats['vital'] = LevelMod[LEVEL]['vital']
base_stats['dh'] = LevelMod[LEVEL]['main']
base_stats['crit'] = LevelMod[LEVEL]['main']
base_stats['det'] = LevelMod[LEVEL]['main']
base_stats['sks'] = LevelMod[LEVEL]['main']
base_stats['tncpt'] = LevelMod[LEVEL]['main']
base_stats['Elemental'] = 3558
base_stats['haste'] = 0

# OUTPUT
expDmg = 0.0
BiS_stats = base_stats.copy()
BiS = {}

##########

# 장비 SET
equips = {}
# 고정 장비
equips['hands'] = item_dict[직군]['손'] # ['+2']
equips['legs'] = item_dict[직군]['다리'] # ['+2']
equips['feet'] = item_dict[직군]['발'] # ['+2']

# for e, value in equips.items():
#     for stats in value.values(): # range(0, len(sum)):
#         for i in range(0, 8): # len(stats)):
#             # if istats[i] in istats:
#                 cellvalue = stats[i]
#                 if cellvalue != None:
#                     base_stats[istats[i]] += cellvalue

# print(base_stats)

# for equip, value in itemset.items():
#     pass


def dh2tnc(statvalues):
    if 직군 == '수호자' or 직군 == '치유사':
        temp = statvalues[eStats['dh']]
        statvalues[eStats['dh']] = None
        statvalues[eStats['tncpt']] = temp

def AddMateria(value, mb, mslot, addval=16, maxval=110) -> bool:
    if value[mb] == None:
        value[mb] = addval
        value[mslot] = istats[mb]
        return True
    elif value[mb] + addval < maxval:
        value[mb] += addval
        value[mslot] = istats[mb]
        return True
    else:
        return False


def calcChain(itemset, part, NextChain):
    for partname, value in itemset[part].items():
        equips[part] = {partname: value}

        NextChain()



# 무기
for weapon, value1 in itemset['weapon'].items():
    if 'weapon' in equips:
        equips['weapon'].clear()
    value1[eStats['dh']] = 80
    value1[eStats['crit']] = 114
    value1[eStats['det']] = 80
    equips['weapon'] = {weapon: value1}

    for head, value2 in itemset['head'].items():
        # if 'head' in equips:
        #     equips['head'].clear()
        equips['head'] = {head: value2}

        for body, value3 in itemset['body'].items():
            # if 'body' in equips:
            #     equips['body'].clear()
            equips['body'] = {body: value3}

            if body != '기린' and body != '주홍':
                continue
            else:
                if body == '주홍':
                    equips['head'].clear()
                for mb1 in range(2,6):
                    value31 = value3.copy()
                    if AddMateria(value31, mb1, 9) == False:
                        continue
                    for mb2 in range(2,6):
                        value32 = value31.copy()
                        if AddMateria(value32, mb2, 10) == False:
                            continue
                        for mb3 in range(2,6):
                            value33 = value32.copy()
                            if AddMateria(value33, mb3, 11) == False:
                                continue
                            for mb4 in range(2,6):
                                value34 = value33.copy()
                                if AddMateria(value34, mb4, 12) == False:
                                    continue
                                for mb5 in range(2,6):
                                    value35 = value34.copy()
                                    if AddMateria(value35, mb5, 13) == False:
                                        continue
                                    equips['body'] = {body: value35}

                                    for earring, value6 in itemset['earrings'].items():
                                        dh2tnc(value6)
                                        # if 'earrings' in equips:
                                        #     equips['earrings'].clear()                
                                        equips['earrings'] = {earring: value6}

                                        for necklace, value7 in itemset['necklace'].items():
                                            dh2tnc(value7)
                                            # if 'necklace' in equips:
                                            #     equips['necklace'].clear()
                                            equips['necklace'] = {necklace: value7}

                                            for bracelet, value8 in itemset['bracelets'].items():
                                                dh2tnc(value8)
                                                # if 'bracelets' in equips:
                                                #     equips['bracelets'].clear()
                                                equips['bracelets'] = {bracelet: value8}

                                                for ring1, value9 in itemset['ring1'].items():
                                                    dh2tnc(value9)
                                                    # if 'ring1' in equips:
                                                    #     equips['ring1'].clear()
                                                    equips['ring1'] = {ring1: value9}

                                                    for ring2, value10 in itemset['ring2'].items():
                                                        dh2tnc(value10)
                                                        # if 'ring2' in equips:
                                                        #     equips['ring2'].clear()
                                                        equips['ring2'] = {ring2: value10}


                                                        tmp_stats = base_stats.copy()

                                                        for name, value in equips.items():
                                                            for stats in value.values(): # range(0, len(sum)):
                                                                for i in range(0, 9): # len(stats)):
                                                                    # if istats[i] in istats:
                                                                        cellvalue = stats[i]
                                                                        if cellvalue != None:
                                                                            tmp_stats[istats[i]] += cellvalue
                                                        
                                                        result = ExpDmgSum(tmp_stats['attr'], tmp_stats['dh'], tmp_stats['crit'],
                                                                        tmp_stats['det'], tmp_stats['sks'], tmp_stats['tncpt'], tmp_stats['haste'])
                                                        
                                                        if result > expDmg:
                                                            print(result)
                                                            expDmg = result
                                                            BiS = equips.copy()
                                                            BiS_stats.update(tmp_stats)
                                                            for id in iequips:
                                                                print(f'{id} : {BiS[id]}')
                                                            print(BiS_stats)



##########

# R E S U L T
# result = ExpDmgSum(sum['attr'], sum['dh'], sum['crit'],
#                    sum['det'], sum['sks'], sum['tncpt'], sum['haste'])
# for ix in iequips:
#     print(f'{ix} : {len(itemset[ix])}')

print('{:8.3f}'.format(expDmg))
print(BiS_stats)
# print(BiS)
    # if id in BiS:
for id in iequips:
        print(f'{id} : {BiS[id]}')
