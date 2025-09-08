import os
import openpyxl
import pandas as pd
from datetime import datetime
import time
import itertools
import math

from util.cdhsp import *
import io
from enum import Enum, auto
from util.xiv import *  # ITEM, ITEMTYPE

start = time.time()

class ItemSet:
    def __init__(self, level, job):
        self.job = job
        self.level = level
        # self.calc = StatsCalc(level)
        self.equips = { x: [] for x in iequips }  # Initialize all equipment slots with empty lists
        pass
    pass


statcalc = StatsCalc(100)
base = ItemSet(100, 'WHM')

TARGET_GCD = 2.50
def need_sks(target_GCD):
    sks = 0
    while statcalc.GCDmod(sks) > target_GCD:
        sks += 1
    return sks
# REQ_SKS = need_sks(TARGET_GCD)

stat_base = {k: 0 for v,k in enumerate(istats)}
stat_base['attr'] = 439  # Base attribute for calculations
stat_base['dh'] = 420  # Base DH for calculations
stat_base['crit'] = 420  # Base Crit for calculations
stat_base['det'] = 440  # Base Det for calculations
stat_base['sks'] = 420  # Base SKS for calculations
stat_base['tncpt'] = 420  # Base Tenacity for calculations
iset = {k: None for v,k in enumerate(base.equips)}
bis = iset.copy()
best_stat = stat_base.copy()
bisMatA = {k: 0 for k in istats[1:6]}  # Materia A
bisMatB = bisMatA.copy()  # Materia B
exp100p = 1.0

### Week 1
TombMax = 900
TombCosts = {'W':900, 'A':825, 'B':495, 'C':375}


### Load the Excel file
path = r'E:\Users\i\Documents\My Games\FINAL FANTASY XIV - KOREA\docs\7황금'
file = r'760items.xlsx'
file_path = os.path.join(path, file)
print(f'\nLoading {file} from {path}...')


def getItemsFromDB(file_path, req_sks=0):
    item_dict = {}

    wb = openpyxl.load_workbook(file_path)
    sheet = wb[wb.sheetnames[0]]  # Assuming the first sheet is the one we want
    ### Item Data
    headers = [cell.value for cell in sheet[1]]  # Get headers from the first row
    ### ['부위', '직군', '획득', '주', '극', '의', '시', '직', '굴/신']
    ItemHeaders = Enum('ItemHeaders', {header.replace('/', '_'): auto() for header in headers})
    temp = {}
    for row in sheet.iter_rows(2, values_only=True):
        if row[0] not in item_dict:
            item_dict[row[0]] = []

        ### JOB 착용가능 type 직군 확인
        #부위
        if row[0] == 'food':
            pass
        elif row[6] != None and req_sks == 0:
            continue
        #직군
        elif row[1] not in [j for j in JOB[base.job]] and row[1] not in [j.value for j in JOB[base.job][1:3]]:
            continue
        ### 획득: 출발 / 최종
        elif row[2] not in ['제작', '일반', '토벌']:
        # elif row[2] not in ['일반', '보강', '레이드']:
            continue
        idata = { headers[i]: row[i] for i in range(2, len(headers)) }
        idata[headers[1]] = row[1] if row[1] is not None else temp[headers[1]]

        item_dict[row[0]].append(idata)
        # item_dict.append(idata)
        temp = idata
        pass
    return item_dict

item_dict = getItemsFromDB(file_path)
# print(f'{base.job} - {len(item_dict)} items.')
print(len(base.equips))

# iequips의 각 슬롯에 대해 item_dict에서 해당 부위의 아이템을 대입
for slot in range(len(iequips)):
    if iequips[slot] == 'ring2':
        base.equips[iequips[slot]] = item_dict[dequips[slot-1]]
    else:
        base.equips[iequips[slot]] = item_dict[dequips[slot]]
    pass

### Calculation


def ProcessSlot(slot_idx, input_stat, matA=None, matB=None):
    global iset
    if slot_idx == len(list(base.equips.keys())):
        # key = list(base.equips.keys())[slot_idx]
        # print(f'AddStat: {input_stat}')
        # CALCULATEMATERIA

        CalculateMateria(input_stat.copy(), matA, matB)
        return


    key = list(base.equips.keys())[slot_idx]
    for s in base.equips[key]:
        if key == 'ring2' and s['획득'] != '제작' and s['획득'] == iset['ring1']['획득']:
            continue
        iset[key] = s
        istat = {
            'attr': int(s['주'] or 0),
            'dh': int(s['직'] or 0),
            'crit': int(s['극'] or 0),
            'det': int(s['의'] or 0),
            'sks': int(s['시'] or 0),
            'tncpt': int(s['굴/신'] or 0),
            'matA': int(s.get('matA', 0) or 0),
            'matB': int(s.get('matB', 0) or 0)
        }

        stat_keys = list(istat.keys())[1:6] #['dh', 'crit', 'det', 'sks', 'tncpt']
        matsA = {k: 0 for k in stat_keys}
        matsB = matsA.copy()

        if slot_idx <= 10:
            # 5개 중 큰 값 2개
            [max, sub] = sorted([istat[k] for k in stat_keys], reverse=True)[:2]

            for skey in stat_keys:
                kvalue = max - istat[skey]
                matsA[skey] = min(istat['matA'], int(kvalue / 54))
                kvalue -= 54 * matsA[skey]
                if istat['matB'] > 0: # 제작금단
                    matsB[skey] = min(istat['matB'], int(kvalue / 18))
                pass
            pass

        if matA is not None:
            matsA = {k: matA.get(k, 0) + matsA.get(k, 0) for k in matsA}
        if matB is not None:
            matsB = {k: matB.get(k, 0) + matsB.get(k, 0) for k in matsB}
        istat = {k: input_stat.get(k, 0) + int(istat.get(k, 0) or 0)
                 for k in input_stat}

        ProcessSlot(slot_idx +1, istat.copy(),
                    matsA.copy(), matsB.copy())
        pass
    pass

def CalculateMateria(input_stat, matA=None, matB=None):
    matsA = {}
    matsB = {}
    for combA in itertools.product(range(0, matA['dh']+1)
                                   , range(0, matA['crit']+1)
                                   , range(0, matA['det']+1)
                                   , range(0, matA['sks']+1)
                                   ):
        if sum(combA) == input_stat['matA']:
            for combB in itertools.product(range(0, matB['dh']+1)
                                           , range(0, matB['crit']+1)
                                           , range(0, matB['det']+1)
                                           , range(0, matB['sks']+1)
                                           ):
                if sum(combB) == input_stat['matB']:
                    sum_stats = input_stat.copy()
                    for key in range(0, len(combA)):
                        sum_stats[istats[1+key]] += combA[key]*54
                        sum_stats[istats[1+key]] += combB[key]*18
                        matsA[istats[1+key]] = combA[key]
                        matsB[istats[1+key]] = combB[key]

                    expGCD = statcalc.GCDmod(sum_stats['sks'])

                    if expGCD != TARGET_GCD:
                        continue
                    # compareBiS(sum_stats, in_equips, matsA, matsB)
                    return compareBiS(sum_stats, matsA, matsB)

def compareBiS(sum_stats, matA, matB):
    global exp100p, bis, bisMatA, bisMatB, iset
    expected = statcalc.ExpDmgMult(sum_stats)
    if expected > exp100p:
        exp100p = expected
        bis.update(iset)
        best_stat.update(sum_stats)
        bisMatA = matA.copy()
        bisMatB = matB.copy()
        # print(f'New BIS: {bis}')
    pass

ProcessSlot(0, stat_base)

end = time.time()
elapsed = end - start
print(f'Time taken: {elapsed:.2f} seconds')

print(f'exp100p: {exp100p:.2f}')
print(f'\nBest Stats: {best_stat}')
print(f'{'부위':8}', end=' ')
for k in istats:
    print(f'{k:4}', end=' ')
for slot in bis:
    print(f'\n{slot:8}:', end=' ')
    # {bis[slot]['획득']} {bis[slot]['주']:4} {bis[slot]['직']:4} {bis[slot]['극']:4} {bis[slot]['의']:4} {bis[slot]['시']:4} {bis[slot]['굴/신']:4}')
    for k in bis[slot]:
        if bis[slot][k] is None:
            print(f'{'-':4}', end=' ')
        else:
            print(f'{bis[slot][k]:4}', end=' ')
print()
print(f'BIS Materia A: {bisMatA}')
print(f'BIS Materia B: {bisMatB}')